from openpyxl import Workbook
import os


total=int(input("Enter the incoming/current Mpesa Amount "))
target=0.4*total
kcb_savings=0.2*total
total_savings=0.6*total

wb=Workbook()
ws=wb.active
ws.title ="My Saving Structure"
ws.append(["Mpesa In","Target Savings","KCB savings","Total Savings"])

ws['A2']=total
ws['B2']=target
ws['c2']=kcb_savings
ws['d2']=total_savings



os.chdir('C:\\Users\\User')

os.system('start excel.exe savings.xlsx')
wb.save("savings.xlsx")